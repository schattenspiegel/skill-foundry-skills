#!/usr/bin/env python3
"""Emit a bounded, read-only semantic and OOXML inventory for an Excel workbook."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Iterable
from pathlib import Path
from typing import Any

SCHEMA_VERSION = 1
DEFAULT_DETAILS_LIMIT = 100
FORMULA_PREFIX = "="
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"

RISK_PATTERNS: tuple[tuple[str, re.Pattern[str]], ...] = (
    ("vba", re.compile(r"(^|/)vbaProject(?:Signature)?\.bin$", re.I)),
    ("activex", re.compile(r"(^|/)(activeX|ctrlProps)/", re.I)),
    ("embeddings", re.compile(r"(^|/)embeddings/", re.I)),
    ("drawings", re.compile(r"(^|/)drawings/", re.I)),
    ("charts", re.compile(r"(^|/)charts/", re.I)),
    ("chartsheets", re.compile(r"(^|/)chartsheets/", re.I)),
    ("pivot_tables", re.compile(r"(^|/)pivotTables/", re.I)),
    ("pivot_caches", re.compile(r"(^|/)pivotCache/", re.I)),
    ("slicers", re.compile(r"(^|/)(slicers|slicerCaches|timelines)/", re.I)),
    ("external_links", re.compile(r"(^|/)externalLinks/", re.I)),
    ("connections", re.compile(r"(^|/)(connections\.xml|queryTables/)", re.I)),
    ("custom_xml", re.compile(r"^customXml/", re.I)),
    ("threaded_comments", re.compile(r"(^|/)(threadedComments|persons)/", re.I)),
    ("comments", re.compile(r"(^|/)comments\d*\.xml$", re.I)),
    ("media", re.compile(r"(^|/)media/", re.I)),
    ("custom_properties", re.compile(r"^docProps/custom\.xml$", re.I)),
    ("signatures", re.compile(r"(^|/)(_xmlsignatures|signatures)/", re.I)),
)

KNOWN_PREFIXES = (
    "[Content_Types].xml",
    "_rels/",
    "docProps/",
    "xl/_rels/",
    "xl/theme/",
    "xl/styles.xml",
    "xl/sharedStrings.xml",
    "xl/workbook.xml",
    "xl/worksheets/",
    "xl/tables/",
    "xl/calcChain.xml",
    "xl/printerSettings/",
    "xl/comments",
    "xl/drawings/",
    "xl/charts/",
    "xl/chartsheets/",
    "xl/pivotTables/",
    "xl/pivotCache/",
    "xl/externalLinks/",
    "xl/connections.xml",
    "xl/queryTables/",
    "xl/media/",
    "xl/embeddings/",
    "xl/activeX/",
    "xl/ctrlProps/",
    "customXml/",
)


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def stable_digest(items: Iterable[object]) -> str:
    payload = json.dumps(list(items), ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def bounded(items: list[Any], limit: int) -> dict[str, object]:
    return {
        "count": len(items),
        "items": items[:limit],
        "truncated": len(items) > limit,
        "digest": stable_digest(items),
    }


def formula_pattern(formula: str, coordinate: str) -> str:
    try:
        from openpyxl.formula.translate import Translator, TranslatorError
    except ImportError:
        return formula
    try:
        return Translator(formula, origin=coordinate).translate_formula("ZZ10000")
    except (TypeError, ValueError, TranslatorError):
        return formula


def package_inventory(path: Path, limit: int) -> dict[str, object]:
    with zipfile.ZipFile(path) as archive:
        names = sorted(info.filename for info in archive.infolist() if not info.is_dir())
        risks: dict[str, list[str]] = {}
        for category, pattern in RISK_PATTERNS:
            matched = [name for name in names if pattern.search(name)]
            if matched:
                risks[category] = matched
        opaque_hashes = {
            name: sha256_bytes(archive.read(name))
            for name in names
            if any(
                category in {"vba", "activex", "embeddings", "signatures"}
                and pattern.search(name)
                for category, pattern in RISK_PATTERNS
            )
        }
        relationships: list[dict[str, str]] = []
        malformed_relationships: list[dict[str, str]] = []
        for name in names:
            if not name.endswith(".rels"):
                continue
            try:
                root = ET.fromstring(archive.read(name))
                for rel in root.findall(f"{{{REL_NS}}}Relationship"):
                    relationships.append(
                        {
                            "part": name,
                            "id": rel.attrib.get("Id", ""),
                            "type": rel.attrib.get("Type", ""),
                            "target": rel.attrib.get("Target", ""),
                            "target_mode": rel.attrib.get("TargetMode", "Internal"),
                        }
                    )
            except (ET.ParseError, KeyError) as exc:
                malformed_relationships.append({"part": name, "error": type(exc).__name__})
        relationships.sort(key=lambda item: (item["part"], item["id"], item["target"]))
        unknown_parts = [name for name in names if not name.startswith(KNOWN_PREFIXES)]
        return {
            "parts": bounded(names, limit),
            "relationships": bounded(relationships, limit),
            "risk_parts": {key: bounded(value, limit) for key, value in sorted(risks.items())},
            "unknown_parts": bounded(unknown_parts, limit),
            "malformed_relationships": malformed_relationships,
            "opaque_hashes": opaque_hashes,
            "vba_present": "xl/vbaProject.bin" in names,
            "vba_sha256": opaque_hashes.get("xl/vbaProject.bin"),
        }


def workbook_inventory(path: Path, limit: int, include_values: bool) -> dict[str, object]:
    from openpyxl import load_workbook

    keep_vba = path.suffix.lower() in {".xlsm", ".xltm"}
    workbook = load_workbook(
        path,
        read_only=False,
        data_only=False,
        keep_vba=keep_vba,
        keep_links=True,
    )
    cached = (
        load_workbook(path, read_only=False, data_only=True, keep_vba=keep_vba, keep_links=True)
        if include_values
        else None
    )
    formula_records: list[dict[str, object]] = []
    value_records: list[dict[str, object]] = []
    sheets: list[dict[str, object]] = []
    for sheet in workbook.worksheets:
        formulas = 0
        nonempty = 0
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is not None:
                    nonempty += 1
                if isinstance(value, str) and value.startswith(FORMULA_PREFIX):
                    formulas += 1
                    record: dict[str, object] = {
                        "sheet": sheet.title,
                        "cell": cell.coordinate,
                        "formula": value,
                        "pattern": formula_pattern(value, cell.coordinate),
                    }
                    if cached is not None:
                        record["cached"] = cached[sheet.title][cell.coordinate].value
                    formula_records.append(record)
                elif include_values and value is not None:
                    value_records.append(
                        {
                            "sheet": sheet.title,
                            "cell": cell.coordinate,
                            "value": value,
                            "data_type": cell.data_type,
                            "number_format": cell.number_format,
                        }
                    )
        tables = [
            {"name": table.name, "display_name": table.displayName, "ref": table.ref}
            for table in sheet.tables.values()
        ]
        tables.sort(key=lambda item: item["name"])
        validations = []
        if sheet.data_validations is not None:
            validations = [
                {"type": item.type, "sqref": str(item.sqref), "formula1": item.formula1}
                for item in sheet.data_validations.dataValidation
            ]
        sheets.append(
            {
                "name": sheet.title,
                "state": sheet.sheet_state,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "nonempty_cells": nonempty,
                "formula_cells": formulas,
                "tables": tables,
                "merged_ranges": sorted(str(item) for item in sheet.merged_cells.ranges),
                "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
                "auto_filter": str(sheet.auto_filter.ref) if sheet.auto_filter.ref else None,
                "data_validations": validations,
                "charts": len(sheet._charts),
                "images": len(sheet._images),
            }
        )
    names = []
    for name, defined_name in workbook.defined_names.items():
        names.append(
            {
                "name": name,
                "attr_text": defined_name.attr_text,
                "hidden": bool(defined_name.hidden),
                "local_sheet_id": defined_name.localSheetId,
            }
        )
    names.sort(key=lambda item: (item["name"], item["local_sheet_id"] or -1))
    calculation = getattr(workbook, "calculation", None)
    calculation_payload = {
        key: getattr(calculation, key, None)
        for key in (
            "calcMode",
            "calcId",
            "fullCalcOnLoad",
            "forceFullCalc",
            "iterate",
            "iterateCount",
            "iterateDelta",
        )
    }
    return {
        "epoch": getattr(workbook.epoch, "isoformat", lambda: str(workbook.epoch))(),
        "iso_dates": bool(workbook.iso_dates),
        "calculation": calculation_payload,
        "sheets": sheets,
        "defined_names": bounded(names, limit),
        "formulas": bounded(formula_records, limit),
        "values": bounded(value_records, limit) if include_values else None,
    }


def inspect(
    path: Path,
    *,
    limit: int = DEFAULT_DETAILS_LIMIT,
    include_values: bool = False,
) -> dict[str, object]:
    package = package_inventory(path, limit)
    workbook = workbook_inventory(path, limit, include_values)
    risk_count = sum(int(value["count"]) for value in package["risk_parts"].values())
    unknown_count = int(package["unknown_parts"]["count"])
    malformed_count = len(package["malformed_relationships"])
    return {
        "schema_version": SCHEMA_VERSION,
        "ok": True,
        "path": str(path),
        "extension": path.suffix.lower(),
        "size": path.stat().st_size,
        "package": package,
        "workbook": workbook,
        "preservation": {
            "risk_part_count": risk_count,
            "unknown_part_count": unknown_count,
            "malformed_relationship_count": malformed_count,
            "requires_escalation": bool(risk_count or unknown_count or malformed_count),
            "note": "Risk inventory is not proof of library preservation support.",
        },
    }


def parser() -> argparse.ArgumentParser:
    result = argparse.ArgumentParser(description=__doc__)
    result.add_argument("workbook", type=Path)
    result.add_argument("--include-values", action="store_true")
    result.add_argument("--details-limit", type=int, default=DEFAULT_DETAILS_LIMIT)
    return result


def main(argv: list[str] | None = None) -> int:
    args = parser().parse_args(argv)
    if args.details_limit < 0:
        print(
            json.dumps(
                {
                    "schema_version": SCHEMA_VERSION,
                    "ok": False,
                    "error": "details-limit must be non-negative",
                }
            )
        )
        return 2
    try:
        report = inspect(
            args.workbook.resolve(),
            limit=args.details_limit,
            include_values=args.include_values,
        )
    except (OSError, ValueError, zipfile.BadZipFile, ET.ParseError, ImportError) as exc:
        print(
            json.dumps(
                {
                    "schema_version": SCHEMA_VERSION,
                    "ok": False,
                    "error": f"{type(exc).__name__}: {exc}",
                }
            )
        )
        return 2
    print(json.dumps(report, indent=2, sort_keys=True, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
